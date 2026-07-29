"""Asset import/export tests (sqlite).

Covers: CSV parsing (attr.<key>, tags, attributes_json), dry-run report,
per-row errors, upsert by each match key, non-destructive semantics, and
CSV/XLSX rendering of filtered exports.

Run: uv run pytest src/testing/test_assets_import_export.py
"""

import io

import pytest
from sqlalchemy import select

from src.api.schemas.assets import AssetCreate
from src.assets import io as assets_io
from src.assets.registry import sync_definitions
from src.assets.service import AssetService
from src.core.orm import AssetORM


@pytest.fixture()
async def session(asset_session_factory):
    async with asset_session_factory() as s:
        await sync_definitions(s)
        yield s


CSV_TEXT = (
    "name,ref,asset_type,serial_number,tags,attr.os_version,attributes_json\r\n"
    'FW One,fw1,firewall,SN-1,edge;prod,7.4.5,"{""ha_mode"": ""standalone""}"\r\n'
    "SW One,sw1,switch,SN-2,,,\r\n"
)


def test_parse_csv_rows():
    rows = assets_io.parse_csv_rows(CSV_TEXT)
    assert rows[0]["name"] == "FW One"
    assert rows[0]["tags"] == ["edge", "prod"]
    assert rows[0]["attributes"] == {"os_version": "7.4.5", "ha_mode": "standalone"}
    assert "attributes" not in rows[1]


async def test_import_dry_run_then_apply(session):
    rows = assets_io.parse_csv_rows(CSV_TEXT)
    report = await assets_io.import_assets(
        session, "t1", rows, match_key="ref", dry_run=True, actor="u",
    )
    assert report.dry_run and report.created == 2 and report.failed == 0
    # dry run writes nothing
    count = len((await session.execute(select(AssetORM))).scalars().all())
    assert count == 0

    report = await assets_io.import_assets(
        session, "t1", rows, match_key="ref", dry_run=False, actor="u",
    )
    assert report.created == 2 and report.updated == 0

    # re-import updates by ref, never duplicates (non-destructive upsert)
    rows2 = assets_io.parse_csv_rows(CSV_TEXT.replace("SN-1", "SN-1-NEW"))
    report = await assets_io.import_assets(
        session, "t1", rows2, match_key="ref", dry_run=False, actor="u",
    )
    assert report.created == 0 and report.updated == 2
    fw = (await session.execute(
        select(AssetORM).where(AssetORM.ref == "fw1"))).scalar_one()
    assert fw.serial_number == "SN-1-NEW"


async def test_import_row_errors(session):
    rows = [
        {"name": "Ok", "ref": "ok1", "asset_type": "generic"},
        {"ref": "missing-name", "asset_type": "generic"},          # no name
        {"name": "BadType", "ref": "bt", "asset_type": "spaceship"},
        {"name": "BadAttr", "ref": "ba", "asset_type": "firewall",
         "attributes": {"ha_mode": "quantum"}},
    ]
    report = await assets_io.import_assets(
        session, "t1", rows, match_key="ref", dry_run=False, actor="u",
    )
    assert report.created == 1 and report.failed == 3
    failures = {r.row: r.errors for r in report.rows if r.action == "error"}
    assert any("name is required" in e for e in failures[2])
    assert any("unknown asset_type" in e for e in failures[3])
    assert any("ha_mode" in e for e in failures[4])


async def test_import_bad_match_key(session):
    with pytest.raises(Exception):
        await assets_io.import_assets(
            session, "t1", [], match_key="password", dry_run=True, actor="u",
        )


async def test_export_csv_and_xlsx(session):
    svc = AssetService(session, gateway=None)
    svc.gateway = None
    await svc.create_asset("t1", AssetCreate(
        name="FW One", ref="fw1", asset_type="firewall", serial_number="SN-1",
        tags=["edge"], attributes={"os_version": "7.4.5"},
    ), "u")

    rows, _ = await svc.list_assets("t1", {}, page=1, page_size=100)
    headers, data = await assets_io.build_export(session, rows)
    assert "serial_number" in headers and "attr.os_version" in headers

    csv_text = assets_io.render_csv(headers, data)
    lines = csv_text.strip().splitlines()
    assert len(lines) == 2
    assert "SN-1" in lines[1] and "7.4.5" in lines[1]

    xlsx_bytes = assets_io.render_xlsx(headers, data)
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(xlsx_bytes), read_only=True)
    ws = wb["assets"]
    grid = [[c.value for c in row] for row in ws.iter_rows()]
    assert grid[0][:2] == ["id", "name"]
    assert "FW One" in grid[1]
